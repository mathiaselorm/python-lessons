import openpyxl as xl
from openpyxl.chart import BarChart, Reference
wb = xl.load_workbook("transaction.xlsx")
sheet = wb['Sheet1']
#cell = sheet.cell(1, 1)

for row in range(2,sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    calculated_price = cell.value * 0.9
    correct_price_cell = sheet.cell(row, 4)
    correct_price_cell = calculated_price
    #corrected_price = correct_price_cell  
   
wb.save('tarnsaction2.xlsx')
    